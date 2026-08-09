import openpyxl, json, re

wb = openpyxl.load_workbook('D:/huanghailan/Desktop/STV-70-01使用登记表&生物样品领用记录 - 无密码版本 -美菱超低温.xlsx', data_only=True)

# Extract all unique position coordinates from 冰箱状态
ws = wb['冰箱状态']
positions = {}
pattern = re.compile(r'^F(\d+)-L(\d+)C(\d+)-(\d+)$')

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=False):
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            m = pattern.match(cell.value)
            if m:
                positions[cell.value] = {
                    'fridge': int(m.group(1)),
                    'layer': int(m.group(2)),
                    'col': int(m.group(3)),
                    'slot': int(m.group(4))
                }

# Extract status: coordinate cells in B-F, status in I-M (offset +7)
status_data = {}
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=20, values_only=False):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and pattern.match(cell.value):
            coord = cell.value
            col_idx = cell.column
            status_col = col_idx + 7  # B->I, C->J, etc.
            if status_col <= 20:
                status_cell = ws.cell(row=cell.row, column=status_col)
                if status_cell.value is not None and status_cell.value != '':
                    status_data[coord] = str(status_cell.value)

# Also check P-T block: coords in P-T, status would be... let me check the layout
# Actually let me also scan columns P-T for coordinate patterns and get status from same cells
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=16, max_col=20, values_only=False):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and pattern.match(cell.value):
            coord = cell.value
            # For P-T block, status might be in a parallel block
            # Let's just record that this coord exists
            if coord not in status_data:
                status_data[coord] = '空'

# Extract layout info
layout = {}
for coord, info in positions.items():
    f = info['fridge']
    if f not in layout:
        layout[f] = {'layers': set(), 'cols': set(), 'slots': set()}
    layout[f]['layers'].add(info['layer'])
    layout[f]['cols'].add(info['col'])
    layout[f]['slots'].add(info['slot'])

print('=== Freezer Layout ===')
for f in sorted(layout.keys()):
    l = layout[f]
    print(f'F{f}: {len(l["layers"])} layers {sorted(l["layers"])}, {len(l["cols"])} cols {sorted(l["cols"])}, {len(l["slots"])} slots {sorted(l["slots"])}')

print(f'\nTotal positions: {len(positions)}')
print(f'\n=== Non-empty Status ===')
for coord in sorted(status_data.keys()):
    s = status_data[coord]
    if s != '空':
        print(f'{coord}: {s}')

# Extract sample entries from 冰箱存取
ws2 = wb['冰箱存取']
samples = []
for row in ws2.iter_rows(min_row=4, max_row=ws2.max_row, min_col=13, max_col=22, values_only=True):
    date_val = row[0]
    person = row[1]
    project = row[2]
    fridge = row[3]
    drawer = row[4]
    positions_used = [v for v in row[5:9] if v is not None]

    if date_val is not None or project is not None:
        samples.append({
            'date': str(date_val).split(' ')[0] if date_val else '',
            'person': str(person) if person else '',
            'project': str(project) if project else '',
            'fridge': str(fridge) if fridge else '',
            'drawer': str(drawer) if drawer else '',
            'positions': len(positions_used)
        })

print(f'\n=== Sample Entries ({len(samples)}) ===')
for s in samples:
    print(s)

# Generate JSON data for the web app
output = {
    'freezers': {},
    'samples': samples,
    'status': status_data
}

for f in sorted(layout.keys()):
    l = layout[f]
    output['freezers'][f'F{f}'] = {
        'layers': sorted(l['layers']),
        'cols': sorted(l['cols']),
        'slots': sorted(l['slots']),
        'maxLayer': max(l['layers']),
        'maxCol': max(l['cols']),
        'maxSlot': max(l['slots'])
    }

with open('C:/Users/huanghailan/WorkBuddy/2026-08-07-11-01-54/freezer_data.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print('\n=== JSON saved to freezer_data.json ===')
print(json.dumps(output['freezers'], ensure_ascii=False, indent=2))
